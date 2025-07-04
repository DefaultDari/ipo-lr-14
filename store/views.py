from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Q
from .models import Product, Category, Manufacturer, Cart, CartItem
import json
from django.http import Http404
from pathlib import Path
import os
import logging
from io import BytesIO
from django.conf import settings
from django.core.mail import send_mail
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Cart, Order, OrderItem
from rest_framework import viewsets, permissions
from .models import Category, Manufacturer, Product, Cart, CartItem
from .serializers import (
    CategorySerializer,
    ManufacturerSerializer,
    ProductSerializer,
    CartSerializer,
    CartItemSerializer
)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [permissions.IsAuthenticated]

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        # Автоматическое назначение создателя товара
        serializer.save(создатель=self.request.user)

class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()  # Add this line
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        # Пользователь видит только свою корзину
        return Cart.objects.filter(пользователь=self.request.user)

class CartItemViewSet(viewsets.ModelViewSet):
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        # Пользователь видит только свои элементы корзины
        return CartItem.objects.filter(корзина__пользователь=self.request.user)
def load_specialties():
    # Load specialties from dump.json
    BASE_DIR = Path(__file__).resolve().parent.parent
    json_file = BASE_DIR / 'dump.json'
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    specialties = []
    for item in data:
        if item['model'] == 'data.specialty':
            specialties.append(item)
    
    return specialties

def spec_list(request):
    specialties = load_specialties()
    return render(request, 'store/spec_list.html', {
        'specialties': specialties,
        'title': 'Список квалификаций'
    })

def spec_detail(request, pk):
    specialties = load_specialties()
    spec = None
    
    for item in specialties:
        if item['pk'] == int(pk):
            spec = item
            break
    
    if not spec:
        raise Http404("Квалификация не найдена")
    
    return render(request, 'store/spec_detail.html', {
        'spec': spec,
        'title': spec['fields']['title']
    })
def index(request):
    """Главная страница"""
    featured_products = Product.objects.filter(количество_на_складе__gt=0)[:8]
    categories = Category.objects.all()
    return render(request, 'store/index.html', {
        'featured_products': featured_products,
        'categories': categories,
    })


def product_list(request):
    """Список товаров с фильтрацией"""
    products = Product.objects.filter(количество_на_складе__gt=0)
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()
    
    # Фильтрация
    category_id = request.GET.get('category')
    manufacturer_id = request.GET.get('manufacturer')
    search_query = request.GET.get('search')
    
    if category_id:
        products = products.filter(категория_id=category_id)
    
    if manufacturer_id:
        products = products.filter(производитель_id=manufacturer_id)
    
    if search_query:
        products = products.filter(
            Q(название__icontains=search_query) | 
            Q(описание__icontains=search_query)
        )
    
    return render(request, 'store/product_list.html', {
        'products': products,
        'categories': categories,
        'manufacturers': manufacturers,
        'selected_category': int(category_id) if category_id else None,
        'selected_manufacturer': int(manufacturer_id) if manufacturer_id else None,
        'search_query': search_query,
    })


def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    related_products = Product.objects.filter(
        категория=product.категория,
        количество_на_складе__gt=0
    ).exclude(id=product.id)[:4]
    
    return render(request, 'store/product_detail.html', {
        'product': product,
        'related_products': related_products,
    })



@login_required
def cart_view(request):
    """Просмотр корзины"""
    cart, created = Cart.objects.get_or_create(пользователь=request.user)
    return render(request, 'store/cart.html', {'cart': cart})

# Обновленные функции для работы с корзиной
@login_required
@require_POST
def add_to_cart(request, product_id):
    """Добавление товара в корзину"""
    try:
        # Получаем товар и запрошенное количество
        product = get_object_or_404(Product, id=product_id)
        quantity = int(request.POST.get('quantity', 1))
        
        # Проверяем доступное количество
        if quantity <= 0:
            messages.error(request, 'Некорректное количество')
            return redirect('product_detail', pk=product_id)
            
        if quantity > product.количество_на_складе:
            messages.error(
                request, 
                f'Недостаточно товара на складе. Доступно: {product.количество_на_складе} шт.'
            )
            return redirect('product_detail', pk=product_id)
        
        # Получаем или создаем корзину пользователя
        cart, created = Cart.objects.get_or_create(пользователь=request.user)
        
        # Ищем существующий элемент корзины для этого товара
        cart_item, item_created = CartItem.objects.get_or_create(
            корзина=cart,
            товар=product,
            defaults={'количество': quantity}
        )
        
        # Если элемент уже существует, обновляем количество
        if not item_created:
            new_quantity = cart_item.количество + quantity
            if new_quantity > product.количество_на_складе:
                messages.error(
                    request, 
                    f'Превышено доступное количество. Доступно: {product.количество_на_складе} шт.'
                )
                return redirect('cart')
                
            cart_item.количество = new_quantity
            cart_item.save()
            messages.success(
                request, 
                f'Количество товара "{product.название}" обновлено в корзине'
            )
        else:
            messages.success(
                request, 
                f'Товар "{product.название}" добавлен в корзину'
            )
            
        return redirect('cart')
        
    except Exception as e:
        messages.error(request, f'Ошибка при добавлении в корзину: {str(e)}')
        return redirect('product_detail', pk=product_id)

@login_required
@require_POST
def update_cart_item(request, item_id):
    """Обновление количества товара в корзине"""
    try:
        # Получаем элемент корзины текущего пользователя
        cart_item = get_object_or_404(
            CartItem, 
            id=item_id, 
            корзина__пользователь=request.user
        )
        
        quantity = int(request.POST.get('quantity', 1))
        
        # Проверяем корректность количества
        if quantity <= 0:
            cart_item.delete()
            messages.success(request, 'Товар удален из корзины')
            return redirect('cart')
            
        # Проверяем доступное количество
        if quantity > cart_item.товар.количество_на_складе:
            messages.error(
                request, 
                f'Недостаточно товара на складе. Доступно: {cart_item.товар.количество_на_складе} шт.'
            )
            return redirect('cart')
        
        # Обновляем количество
        cart_item.количество = quantity
        cart_item.save()
        messages.success(request, 'Количество товара обновлено')
        
        return redirect('cart')
        
    except Exception as e:
        messages.error(request, f'Ошибка при обновлении корзины: {str(e)}')
        return redirect('cart')

@login_required
@require_POST
def remove_from_cart(request, item_id):
    """Удаление товара из корзины"""
    try:
        # Получаем элемент корзины текущего пользователя
        cart_item = get_object_or_404(
            CartItem, 
            id=item_id, 
            корзина__пользователь=request.user
        )
        
        product_name = cart_item.товар.название
        cart_item.delete()
        
        messages.success(
            request, 
            f'Товар "{product_name}" удален из корзины'
        )
        
        return redirect('cart')
        
    except Exception as e:
        messages.error(request, f'Ошибка при удалении товара: {str(e)}')
        return redirect('cart')

def register(request):
    """Регистрация пользователя"""
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Регистрация прошла успешно!')
            return redirect('index')
    else:
        form = UserCreationForm()
    
    return render(request, 'registration/register.html', {'form': form})
logger = logging.getLogger(__name__)

@login_required
def checkout(request):
    """Оформление заказа"""
    cart = Cart.objects.get(пользователь=request.user)
    
    if request.method == 'POST':
        try:
            # Создаем заказ
            order = Order.objects.create(
                пользователь=request.user,
                общая_стоимость=cart.общая_стоимость(),
                адрес_доставки=request.POST.get('address'),
                email=request.POST.get('email'),
                телефон=request.POST.get('phone')
            )
            
            # Создаем элементы заказа
            for item in cart.cartitem_set.all():
                OrderItem.objects.create(
                    заказ=order,
                    товар=item.товар,
                    количество=item.количество,
                    цена=item.товар.цена
                )
                # Уменьшаем количество товара на складе
                product = item.товар
                product.количество_на_складе -= item.количество
                product.save()
            
            # Генерируем чек в Excel
            excel_file = generate_excel_receipt(order)
            
            # Отправляем email с чеком
            send_receipt_email(order, excel_file)
            
            # Очищаем корзину
            cart.cartitem_set.all().delete()
            
            messages.success(request, 'Заказ успешно оформлен! Чек отправлен на вашу почту.')
            return redirect('index')
        
        except Exception as e:
            logger.error(f"Ошибка при оформлении заказа: {str(e)}")
            messages.error(request, f'Ошибка при оформлении заказа: {str(e)}')
            return redirect('cart')
    
    return render(request, 'store/checkout.html', {'cart': cart})

def generate_excel_receipt(order):
    """Генерация чека в формате Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Чек заказа"
    
    # Заголовки
    headers = ["Товар", "Количество", "Цена за шт.", "Сумма"]
    ws.append(headers)
    
    # Стиль для заголовков
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    
    # Данные заказа
    for item in order.orderitem_set.all():
        ws.append([
            item.товар.название,
            item.количество,
            item.цена,
            item.количество * item.цена
        ])
    
    # Итог
    ws.append(["", "", "Итого:", order.общая_стоимость])
    ws[f'D{ws.max_row}'].font = bold_font
    
    # Выравнивание
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    
    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def send_receipt_email(order, excel_file):
    """Отправка чека по электронной почте"""
    subject = f"Чек заказа #{order.id}"
    message = (
        f"Спасибо за ваш заказ!\n\n"
        f"Номер заказа: #{order.id}\n"
        f"Дата заказа: {order.дата_создания.strftime('%d.%m.%Y %H:%M')}\n"
        f"Общая стоимость: {order.общая_стоимость} ₽\n\n"
        f"Адрес доставки: {order.адрес_доставки}\n"
    )
    
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [order.email],
        fail_silently=False,
        html_message=message.replace('\n', '<br>'),
        attachments=[
            (f"чек_заказа_{order.id}.xlsx", excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        ]
    )